"""
Order Fulfillment and Logistics Exception Tracker
----------------------------------------------------
Builds an Excel operations tracker for open orders, shipment delays,
fulfillment cycle time, aging backorders, shortage risks, and logistics
follow-up actions - the kind of tracker used in daily ops/logistics
stand-ups.

Run: python build_excel_tracker.py
Requires: pandas, openpyxl (pip install pandas openpyxl)
Output: ../Order_Fulfillment_Tracker.xlsx"
"""
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

DATA_PATH = "../data/order_fulfillment_data.csv"
OUT_PATH = "../Order_Fulfillment_Tracker.xlsx"

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
RED_FILL = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")


def build_summary(df: pd.DataFrame) -> pd.DataFrame:
    summary = (
        df.groupby("status")
        .agg(order_count=("order_id", "count"), avg_cycle_time_days=("cycle_time_days", "mean"))
        .reset_index()
        .sort_values("order_count", ascending=False)
    )
    return summary


def build_exception_view(df: pd.DataFrame) -> pd.DataFrame:
    exceptions = df[
        (df["status"] != "Delivered On Time")
    ].sort_values(["backorder_flag", "delay_days"], ascending=[False, False])
    return exceptions


def style_sheet(ws, df: pd.DataFrame) -> None:
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(col_name) + 2)

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.freeze_panes = "A2"


def add_conditional_formatting(ws, df: pd.DataFrame) -> None:
    if "status" not in df.columns:
        return
    status_col_idx = list(df.columns).index("status") + 1
    col_letter = get_column_letter(status_col_idx)
    last_row = len(df) + 1
    rng = f"{col_letter}2:{col_letter}{last_row}"

    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"Backordered"'], fill=RED_FILL)
    )
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"Delivered Late"'], fill=YELLOW_FILL)
    )
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"Partial Shipment"'], fill=YELLOW_FILL)
    )


def main():
    df = pd.read_csv(DATA_PATH)

    summary = build_summary(df)
    exceptions = build_exception_view(df)

    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="All Orders", index=False)
        exceptions.to_excel(writer, sheet_name="Exceptions", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)

        style_sheet(writer.sheets["All Orders"], df)
        style_sheet(writer.sheets["Exceptions"], exceptions)
        style_sheet(writer.sheets["Summary"], summary)

        add_conditional_formatting(writer.sheets["All Orders"], df)
        add_conditional_formatting(writer.sheets["Exceptions"], exceptions)

    print(f"Wrote tracker workbook to {OUT_PATH}")
    print(f"Total orders: {len(df)} | Exceptions: {len(exceptions)} ({len(exceptions)/len(df):.1%})")


if __name__ == "__main__":
    main()
